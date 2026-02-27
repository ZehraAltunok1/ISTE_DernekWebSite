
"""
Excel Report Generator for Dernek Platform
Uses openpyxl to create formatted Excel files
"""

import sys
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_excel_report(data, output_path):
    """
    Create formatted Excel file from data array
    
    Args:
        data: List of lists (rows and columns)
        output_path: Path to save the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Rapor"
    

    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0070F2', end_color='0070F2', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    data_font = Font(name='Arial', size=10)
    data_alignment = Alignment(horizontal='left', vertical='center')
    
    border_thin = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    

    for row_idx, row_data in enumerate(data, start=1):
        for col_idx, cell_value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            
        
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border_thin
            else:
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border_thin
    

    for col_idx in range(1, len(data[0]) + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        
        for row_idx in range(1, len(data) + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        
   
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
 
    ws.freeze_panes = 'A2'
    

    wb.save(output_path)
    print(f"Excel file created: {output_path}")

def main():
    if len(sys.argv) < 3:
        print("Usage: python3 generate_excel.py <json_data> <output_path>")
        sys.exit(1)
    
    try:
   
        json_data = sys.argv[1]
        output_path = sys.argv[2]
        
     
        data = json.loads(json_data)
        
        if not isinstance(data, list):
            raise ValueError("Data must be a list of lists")
        
        
        create_excel_report(data, output_path)
        sys.exit(0)
        
    except Exception as e:
        print(f"Error: {str(e)}", file=sys.stderr)
        sys.exit(1)

if __name__ == "__main__":
    main()